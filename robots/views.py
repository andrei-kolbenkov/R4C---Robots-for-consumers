from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font

from .forms import RobotForm
from .models import Robot

@csrf_exempt  # Отключение проверки CSRF (только для API)
def create_robot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Считываем JSON из тела запроса
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        form = RobotForm(data)
        if form.is_valid():
            robot = form.save()
            return JsonResponse({
                'id': robot.id,
                'model': robot.model,
                'version': robot.version,
                'created': robot.created
            }, status=201)
        else:
            return JsonResponse({'errors': form.errors}, status=400)
    return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)


def export_weekly_summary(request):
    # Определяем временной диапазон
    now = datetime.now()
    week_ago = now - timedelta(days=7)

    # Группируем данные по модели и версии
    summary = Robot.objects.filter(created__range=[week_ago, now]).values('model', 'version').annotate(total=Count('id'))

    # Создаём новый Excel-файл
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active  # Получаем стандартный лист

    # Создаём листы для каждой модели
    models = summary.values_list('model', flat=True).distinct()
    if not models:
        # Если нет данных за неделю, заполняем стандартный лист сообщением
        default_sheet.title = "Summary"
        default_sheet.append(["Нет данных за последние 7 дней"])
    else:
        for model in models:
            worksheet = workbook.create_sheet(title=model)

            # Добавляем заголовок
            worksheet.append(['Модель', 'Версия', 'Количество за неделю'])
            worksheet['A1'].font = worksheet['B1'].font = worksheet['C1'].font = Font(bold=True)

            # Добавляем строки для каждой версии
            for row in summary.filter(model=model):
                worksheet.append([row['model'], row['version'], row['total']])

        # Удаляем стандартный лист только если созданы другие листы
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(default_sheet)

    # Генерируем ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="weekly_summary_{now.strftime("%Y%m%d")}.xlsx"'

    workbook.save(response)
    return response